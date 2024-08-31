# from flask import Flask, request, jsonify
# from transformers import BertTokenizerFast, AutoModelForTokenClassification, AutoConfig
# import torch
# from datetime import datetime
# import logging
# import pandas as pd
# import os
# from openpyxl import load_workbook
# from openpyxl.styles import Alignment
# import time

# app = Flask(__name__)

# # Load the model and tokenizer at the start of the application
# # config = AutoConfig.from_pretrained(r'C:\Users\admin\Desktop\Data science\vs\aunty\render\llm_android\LLM_Android\model\config.json')
# # model = AutoModelForTokenClassification.from_pretrained(r'C:\Users\admin\Desktop\Data science\vs\aunty\render\llm_android\LLM_Android\model')
# # tokenizer = BertTokenizerFast.from_pretrained(r'C:\Users\admin\Desktop\Data science\vs\aunty\render\llm_android\LLM_Android\tokens')

# config = AutoConfig.from_pretrained('model/config.json')
# model = AutoModelForTokenClassification.from_pretrained('model')
# tokenizer = BertTokenizerFast.from_pretrained('tokens')

# DEFAULT_GROUPS = [ 'Mobile','Braodband','TDS', 'Salary',"Mobile Payment", 'Biowaste','Investment and Deposits','Loan', 'Rent', 'EB','UPI Payment', 'OTT','Swiggy','Others']

# cumulative_results = {
#     "details": [],
#     "date": [],
#     "Group": []
# }

# # Temporary in-memory storage for data when local drive is not accessible
# temp_storage = []

# EXCEL_PATH = r'C:\Users\admin\Desktop\AUNTY\android.xlsx'


# def extract_year(date):
#     return pd.to_datetime(date).strftime('%Y')

# def extract_month(date):
#     return pd.to_datetime(date).strftime('%B')

# def create_new_sheet(writer, year):
#     df_new = pd.DataFrame({'Group': DEFAULT_GROUPS})
#     for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#         df_new[month + '_Details'] = ''
#         df_new[month + '_Date'] = ''
#     df_new.to_excel(writer, sheet_name=year, index=False)

    

# def save_to_excel(cumulative_results):
#     # Convert the cumulative results to a DataFrame
#     df = pd.DataFrame(cumulative_results)
#     print("Entered save to excel")

#     # Extract year and month from the Date column
#     df['Year'] = df['date'].apply(extract_year)
#     df['Month'] = df['date'].apply(extract_month)
#     df['Month_Details'] = df.apply(lambda x: f"{x['Month']}_Details", axis=1)
#     df['Month_Date'] = df.apply(lambda x: f"{x['Month']}_Date", axis=1)

#     # Assign values to new columns
#     df[df['Month_Details']] = df['details']
#     df[df['Month_Date']] = df['date']
#     print(df)

#     # Check if the file exists
#     if os.path.exists(EXCEL_PATH):
#         print("path exists")
#         # Load the existing Excel file
#         with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#             workbook = writer.book
           

#             for year in df['Year'].unique():
#                 if year not in workbook.sheetnames:
#                     create_new_sheet(writer, year)
                   

#             for year, year_group in df.groupby('Year'):
#                 df_excel = pd.read_excel(EXCEL_PATH, sheet_name=year)

#                 for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                     df_excel[month + '_Details'] = df_excel[month + '_Details'].fillna('').astype(str)
#                     df_excel[month + '_Date'] = df_excel[month + '_Date'].fillna('').astype(str)


#                 for i, row in year_group.iterrows():
                    
#                     columns = ['January_Details','January_Date','February_Details','February_Date','March_Details','March_Date','April_Details','April_Date','May_Details','May_Date','June_Details','June_Date','July_Details','July_Date','August_Details','August_Date','September_Details','September_Date','October_Details','October_Date','November_Details','November_Date','December_Details','December_Date']
                   
#                     row['date'] = pd.to_datetime(row['date'])
                    
#                     month = row['date'].strftime('%B')
#                     row = pd.DataFrame(row).transpose()
#                     row = row.dropna(axis=1,how = 'all')                  
#                     col_details = f"{month}_Details"
#                     col_date = f"{month}_Date"            
#                     details = row[col_details][0]
#                     print("Details",details)
#                     date = row[col_date][0]
#                     print("row",row['Group'].values)
#                     print("excel",df_excel['Group'].values)
                    

                    
#                     matching_idx = df_excel[df_excel['Group'].values == row['Group'].values].index
#                     print("matchning",matching_idx)

#                     if not matching_idx.empty:
                                           
#                         # Update the corresponding row in the Excel DataFrame
#                         idx = matching_idx[0]
#                         print("Idx",df_excel.iloc[:idx+1])
                        
#                         # Create a new row for the new details and date
#                         new_row = {'Group': '', month + '_Details': details, month + '_Date': date}
#                         print("new row",new_row)
#                         df_excel = pd.concat([df_excel.iloc[:idx+1], pd.DataFrame([new_row]), df_excel.iloc[idx+1:]]).reset_index(drop=True)
#                         print("last",df_excel)
#                         # Merge the Group value for the first row
#                         # df_excel.at[idx, 'Group'] = row['Group']
#                     else:    
#                         print("entered else")                   
#                         # If the group does not have an entry, add a new row
#                         new_row = {'Group': row['Group'], month + '_Details': details, month + '_Date': date}
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)
                        
#                 for group in DEFAULT_GROUPS:
                   
#                     if group not in df_excel['Group'].values:
#                         print('entered if loop')   
#                         new_row = {'Group': group}
#                         for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                             new_row[month + '_Details'] = ''
#                             new_row[month + '_Date'] = ''
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)
                
#                 df_excel.to_excel(writer, sheet_name=year, index=False, header=True)
#                 cumulative_results.clear()
#     else:
#         print("path doent exists")
#         # Create a new Excel file with a sheet for each year in the cumulative results
#         with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
#             for year, year_group in df.groupby('Year'):
#                 create_new_sheet(writer, year)
#                 df_excel = pd.DataFrame(columns=['Group'])

#                 for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                     df_excel[month] = [''] * len(DEFAULT_GROUPS)
#                     df_excel[month + '_Date'] = [''] * len(DEFAULT_GROUPS)

#                 for i, row in year_group.iterrows():
#                     group = row['Group']
#                     details = str(row['details'])
#                     date = str(row['date'])
#                     month = row['Month']

#                     group_idx = df_excel[df_excel['Group'] == group].index

#                     if not group_idx.empty:
#                         idx = group_idx[0]
#                         df_excel.at[idx, month + '_Details'] += f"{details}\n"
#                         df_excel.at[idx, month + '_Date'] += f"{date}\n"
#                     else:
#                         new_row = {'Group': group, month + '_Details': details, month + '_Date': date}
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

#                 for group in DEFAULT_GROUPS:
#                     if group not in df_excel['Group'].values:
#                         new_row = {'Group': group}
#                         for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                             new_row[month + '_Details'] = ''
#                             new_row[month + '_Date'] = ''
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

#                 df_excel.to_excel(writer, sheet_name=year, index=False, header=True)
#                 cumulative_results.clear()


# def alternate_save_to_excel(cumulative_results):
   
#     # Convert the cumulative results to a DataFrame
#     df = pd.DataFrame(cumulative_results)
#     print("Entered save to excel")

#     # Extract year and month from the Date column
#     df['Year'] = df['date'].apply(extract_year)
#     df['Month'] = df['date'].apply(extract_month)
#     df['Month_Details'] = df.apply(lambda x: f"{x['Month']}_Details", axis=1)
#     df['Month_Date'] = df.apply(lambda x: f"{x['Month']}_Date", axis=1)

#     # Assign values to new columns
#     df[df['Month_Details']] = df['details']
#     df[df['Month_Date']] = df['date']
#     print(df)

#     # Check if the file exists
#     if os.path.exists(EXCEL_PATH):
#         print("path exists")
#         # Load the existing Excel file
#         with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#             workbook = writer.book
            
#             for year in df['Year'].unique():
#                 if year not in workbook.sheetnames:
#                     create_new_sheet(writer, year)
                   
            
#             for year, year_group in df.groupby('Year'):
            
#                 df_excel = pd.read_excel(EXCEL_PATH, sheet_name=year)
               
#                 for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                     df_excel[month + '_Details'] = df_excel[month + '_Details'].fillna('').astype(str)
#                     df_excel[month + '_Date'] = df_excel[month + '_Date'].fillna('').astype(str)


#                 for i, row in year_group.iterrows():
                    
#                     columns = ['January_Details','January_Date','February_Details','February_Date','March_Details','March_Date','April_Details','April_Date','May_Details','May_Date','June_Details','June_Date','July_Details','July_Date','August_Details','August_Date','September_Details','September_Date','October_Details','October_Date','November_Details','November_Date','December_Details','December_Date']
                   
#                     row['date'] = pd.to_datetime(row['date'])
                    
#                     month = row['date'].strftime('%B')
#                     row = pd.DataFrame(row).transpose()
#                     row = row.dropna(axis=1,how = 'all')                  
#                     col_details = f"{month}_Details"
#                     col_date = f"{month}_Date"            
#                     details = row[col_details][0]
#                     print("Details",details)
#                     date = row[col_date][0]
#                     print("row",row['Group'].values)
#                     print("excel",df_excel['Group'].values)
                    

                    
#                     matching_idx = df_excel[df_excel['Group'].values == row['Group'].values].index
#                     print("matchning",matching_idx)

#                     if not matching_idx.empty:
                                           
#                         # Update the corresponding row in the Excel DataFrame
#                         idx = matching_idx[0]
#                         print("Idx",df_excel.iloc[:idx+1])
                        
#                         # Create a new row for the new details and date
#                         new_row = {'Group': '', month + '_Details': details, month + '_Date': date}
#                         print("new row",new_row)
#                         df_excel = pd.concat([df_excel.iloc[:idx+1], pd.DataFrame([new_row]), df_excel.iloc[idx+1:]]).reset_index(drop=True)
#                         print("last",df_excel)
#                         # Merge the Group value for the first row
#                         # df_excel.at[idx, 'Group'] = row['Group']
#                     else:    
#                         print("entered else")                   
#                         # If the group does not have an entry, add a new row
#                         new_row = {'Group': row['Group'], month + '_Details': details, month + '_Date': date}
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)
                        
#                 for group in DEFAULT_GROUPS:
                   
#                     if group not in df_excel['Group'].values:
#                         print('entered if loop')   
#                         new_row = {'Group': group}
#                         for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                             new_row[month + '_Details'] = ''
#                             new_row[month + '_Date'] = ''
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

#                 df_excel.to_excel(writer, sheet_name=year, index=False, header=True)
              
                
#     else:
#         print("entered else loop")
#         # Create a new Excel file with a sheet for each year in the cumulative results
#         with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
#             for year, year_group in df.groupby('Year'):
#                 create_new_sheet(writer, year)
#                 df_excel = pd.DataFrame(columns=['Group'])

#                 for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                     df_excel[month + '_Details'] = [''] * len(DEFAULT_GROUPS)
#                     df_excel[month + '_Date'] = [''] * len(DEFAULT_GROUPS)

#                 for i, row in year_group.iterrows():
#                     group = row['Group']
#                     details = str(row['details'])
#                     date = str(row['date'])
#                     month = row['Month']
#                     print("details type",type(details))
#                     print("date type",type(date))

#                     group_idx = df_excel[df_excel['Group'] == group].index

#                     if not group_idx.empty:
#                         idx = group_idx[0]
#                         df_excel.at[idx, month] += f"{details}\n"
#                         df_excel.at[idx, month + '_Date'] += f"{date}\n"
#                     else:
#                         new_row = {'Group': group, month: details, month + '_Date': date}
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

#                 for group in DEFAULT_GROUPS:
#                     if group not in df_excel['Group'].values:
#                         new_row = {'Group': group}
#                         for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
#                             new_row[month + '_Details'] = ''
#                             new_row[month + '_Date'] = ''
#                         df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

#                 df_excel.to_excel(writer, sheet_name=year, index=False, header=True)
                

                
                   


# # Function to flush temp_storage
# def flush_temp_storage():
#     global temp_storage
#     print("temp storage",type(temp_storage))
#     success = True
#     for result in temp_storage:
#         try:
#             print("trying alternate function")
#             alternate_save_to_excel(result)
#         except Exception as e:
#             print(f"Failed to save to Excel using alternate method: {e}")
#             success = False
#             break

#     if success:
#         temp_storage = []  # Clear temp_storage only if all operations succeed

# # Decorator to conditionally flush temp_storage before handling requests
# def flush_temp_storage_before_request(func):
#     def wrapper(*args, **kwargs):
#         global temp_storage
#         if temp_storage:
#             flush_temp_storage()
#         return func(*args, **kwargs)
#     return wrapper

# # Apply the decorator to handle_before_request
# @app.before_request
# @flush_temp_storage_before_request
# def handle_before_request():
#     pass

# @app.route('/endpoint', methods=['POST'])
# def endpoint():
#     print("Entered")
#     if request.method == 'POST':
#         data = request.get_json()
#         transaction_details = data.get('transactionDetails')
#         formatted_date = data.get('formattedDate')

#         # Perform inference using the model
#         result = model_inference(transaction_details, formatted_date)
#         cumulative_results = result
#         android_result = {"details": result["details"],"date": result["date"]}
                
#         try:
#             save_to_excel(cumulative_results)
#         except Exception as e:
#             print(f"Failed to save to Excel: {e}")
#             temp_storage.append(cumulative_results)

#         return jsonify(android_result)


# def model_inference(transaction_details, formatted_date):
#     # Tokenize the input text
#     print("transaction det",transaction_details)
#     inputs = tokenizer(transaction_details, return_tensors="pt", truncation=True, padding=True)
#     # Perform model inference
#     with torch.no_grad():
#         outputs = model(**inputs)
       
#     # Process model outputs
#     logits = outputs.logits
#     predicted_label_ids = torch.argmax(logits, dim=2)
#     label_list = ["O", "B-NAME", "I-NAME", "B-SENT_TO", "I-SENT_TO"]

#     # Convert label IDs to label names
#     predictions = [label_list[label_id] for label_id in predicted_label_ids[0].cpu().numpy()]
    
#     # Example: Process predictions (print or log them)
#     tokens = tokenizer.convert_ids_to_tokens(inputs["input_ids"][0])
#     word_pieces = []
#     current_label = "O"
#     collected_sent_to = []

#     for token, prediction in zip(tokens, predictions):
#         if token.startswith("##"):
#             word_pieces.append(token[2:])
#         else:
#             if word_pieces and current_label == "B-SENT_TO":
#                 collected_sent_to.append(''.join(word_pieces))
            
#             if prediction == "B-SENT_TO":
#                 word_pieces = [token]
#             elif prediction == "I-SENT_TO" and word_pieces:
#                 word_pieces.append(token)
#             else:
#                 word_pieces = []

#             current_label = prediction
#     print("current",collected_sent_to)

#     if word_pieces and current_label == "B-SENT_TO":
#         collected_sent_to.append(''.join(word_pieces))
    
#     iteration = collected_sent_to
   
#     mobile = ["AirtelPostpaid","AirtelPostpaidB","jiomobili","jio","jiomobile","viposvf","vipos","vi","airtel"]
#     broad_band = ['AirtelBroadband']
#     ott = ["spotify","netflix",]
#     mobile_payment = ["EDC"]
#     investment = ["ICICI PRUDENTIAL ASSET MA","Lic Of India","AXISDIRECT", "DEPOSIT", "INVESTM", "KOTAK MF", "ICIPRU", "FTMF", "MAX LIFE INSURANCE", "MAX LIFE","SIPPRM"]
#     salary = ["Dhanamma A M","Sumathi","Nagamani","Gayatri", "Sathya Priya S","sridevinagesh","Nagalakshmi Murugan","salary"]
#     biowaste = ["Teknothermindustries"]
#     loan = ['amazeloan']
#     rent = ["Karthik vijay shah"]
#     swiggy = ['swiggy']

#     mobile_lower = [m.lower() for m in mobile]
#     broad_band_lower = [b.lower() for b in broad_band]
#     mobile_payment_lower = [mp.lower() for mp in mobile_payment]
#     ott_lower = [o.lower() for o in ott]
#     investment_lower  = [i.lower() for i in investment]
#     salary_lower  = [s.lower() for s in salary]    
#     biowaste_lower = [b.lower() for b in biowaste]
#     loan_lower = [l.lower() for l in loan]
#     rent_lower = [r.lower() for r in rent]
#     swiggy_lower = [s.lower() for s in swiggy]

    
    
#     # specific_keywords = [,"SriMeenatchiPharma","Sri Sairam Agencies","ASCENT THERAPEUTICS","SUCESS PHARMAAVACCINE",,"INTERNET TAX PAYMENT","Freedom diagonostics","citypharma","LPG SUBSIDY","ANBU PHARMACY","grocery",  "ACVERIFY",]

#     for collected_sent in iteration:
#         collected_sent_to_lower = collected_sent.lower()
#         if any(m in collected_sent_to_lower for m in mobile_lower):
#             group_category = "Mobile"
#             break
#         elif any(b in collected_sent_to_lower for b in broad_band_lower):
#             group_category = "Broadband"
#             break
#         elif any(mp in collected_sent_to_lower for mp in mobile_payment_lower):
#             group_category = "UPI Payment"
#             break
#         elif any(o in collected_sent_to_lower for o in ott_lower):
#             group_category = "OTT"
#             break
#         elif any(i in collected_sent_to_lower for i in investment_lower):
#             group_category = "Investment and Deposits"
#             break
#         elif any(s in collected_sent_to_lower for s in salary_lower):
#             group_category = "Salary"
#             break
#         elif any(b in collected_sent_to_lower for b in biowaste_lower):
#             group_category = "Biowaste"
#             break
#         elif any(l in collected_sent_to_lower for l in loan_lower):
#             group_category = "Loan"
#             break
#         elif any(r in collected_sent_to_lower for r in rent_lower):
#             group_category = "Rent"
#             break
#         elif any(s in collected_sent_to_lower for s in swiggy_lower):
#             group_category = "Swiggy"
#             break
#         else:
#             group_category = "Others"

#     # Prepare the result to be returned
#     result = {
#         "Group": group_category,
#         "details": collected_sent_to,
#         "date": formatted_date,
#     }
#     print("result", result)
#     return result

# if __name__ == '__main__':
#     app.run(debug=True, host='0.0.0.0', port=5000)


import os
from flask import Flask, request, jsonify
import pandas as pd
from github import Github
from io import BytesIO
import openpyxl
from datetime import datetime
import torch
from transformers import AutoConfig, AutoModelForTokenClassification, BertTokenizerFast

app = Flask(__name__)

# GitHub setup
GITHUB_TOKEN = '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'
# GITHUB_TOKEN = os.environ.get('GITHUB_TOKEN')
REPO_NAME = 'vinothvikas1987/LLM_Android'
FILE_PATH = 'android.xlsx'

g = Github(GITHUB_TOKEN)
repo = g.get_repo(REPO_NAME)

# Load the model and tokenizer
config = AutoConfig.from_pretrained('model/config.json')
model = AutoModelForTokenClassification.from_pretrained('model')
tokenizer = BertTokenizerFast.from_pretrained('tokens')
#local
# config = AutoConfig.from_pretrained(r'C:\Users\admin\Desktop\Data science\vs\aunty\render\llm_android\LLM_Android\model\config.json')
# model = AutoModelForTokenClassification.from_pretrained(r'C:\Users\admin\Desktop\Data science\vs\aunty\render\llm_android\LLM_Android\model')
# tokenizer = BertTokenizerFast.from_pretrained(r'C:\Users\admin\Desktop\Data science\vs\aunty\render\llm_android\LLM_Android\tokens')


DEFAULT_GROUPS = ['Mobile', 'Broadband', 'TDS', 'Salary', "Mobile Payment", 'Biowaste', 'Investment and Deposits', 'Loan', 'Rent', 'EB', 'UPI Payment', 'OTT', 'Swiggy', 'Others']

def extract_year(date):
    return pd.to_datetime(date).strftime('%Y')

def extract_month(date): 
    return pd.to_datetime(date).strftime('%B')

def create_new_sheet(writer, year):
    df_new = pd.DataFrame({'Group': DEFAULT_GROUPS})
    for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
        df_new[month + '_Details'] = ''
        df_new[month + '_Date'] = ''
    df_new.to_excel(writer, sheet_name=year, index=False)

def get_excel_from_github():
    file_content = repo.get_contents(FILE_PATH)
    return BytesIO(file_content.decoded_content)

def save_excel_to_github(excel_buffer):
    # repo.update_file(FILE_PATH, f"Update Excel file {datetime.now()}", excel_buffer.getvalue(), repo.get_contents(FILE_PATH).sha)
    excel_buffer.seek(0)
    content = excel_buffer.getvalue()
    
    # Get the current commit
    ref = repo.get_git_ref('heads/main')
    commit = repo.get_commit(ref.object.sha)
    base_tree = commit.tree

    # Create a new blob with the updated Excel file
    blob = repo.create_git_blob(base64.b64encode(content).decode(), "base64")
    element = InputGitTreeElement(path=FILE_PATH, mode='100644', type='blob', sha=blob.sha)

    # Create a new tree with the updated file
    tree = repo.create_git_tree([element], base_tree)

    # Create a new commit
    parent = repo.get_git_commit(ref.object.sha)
    commit = repo.create_git_commit(f"Update Excel file {datetime.now()}", tree, [parent])

    # Update the reference
    ref.edit(commit.sha)

def save_to_excel(cumulative_results):
    df = pd.DataFrame(cumulative_results)
    print("Entered save to excel")

    df['Year'] = df['date'].apply(extract_year)
    df['Month'] = df['date'].apply(extract_month)
    df['Month_Details'] = df.apply(lambda x: f"{x['Month']}_Details", axis=1)
    df['Month_Date'] = df.apply(lambda x: f"{x['Month']}_Date", axis=1)

    df[df['Month_Details']] = df['details']
    df[df['Month_Date']] = df['date']

    excel_buffer = get_excel_from_github()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        workbook = writer.book

        for year in df['Year'].unique():
            if year not in workbook.sheetnames:
                create_new_sheet(writer, year)

        for year, year_group in df.groupby('Year'):
            df_excel = pd.read_excel(excel_buffer, sheet_name=year)

            for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
                df_excel[month + '_Details'] = df_excel[month + '_Details'].fillna('').astype(str)
                df_excel[month + '_Date'] = df_excel[month + '_Date'].fillna('').astype(str)

            for i, row in year_group.iterrows():
                row['date'] = pd.to_datetime(row['date'])
                month = row['date'].strftime('%B')
                row = pd.DataFrame(row).transpose()
                row = row.dropna(axis=1, how='all')
                col_details = f"{month}_Details"
                col_date = f"{month}_Date"
                details = row[col_details][0]
                date = row[col_date][0]

                matching_idx = df_excel[df_excel['Group'].values == row['Group'].values].index

                if not matching_idx.empty:
                    idx = matching_idx[0]
                    new_row = {'Group': '', month + '_Details': details, month + '_Date': date}
                    df_excel = pd.concat([df_excel.iloc[:idx+1], pd.DataFrame([new_row]), df_excel.iloc[idx+1:]]).reset_index(drop=True)
                else:
                    new_row = {'Group': row['Group'], month + '_Details': details, month + '_Date': date}
                    df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

            for group in DEFAULT_GROUPS:
                if group not in df_excel['Group'].values:
                    new_row = {'Group': group}
                    for month in pd.date_range(start='1/1/2024', end='12/31/2024', freq='MS').strftime("%B"):
                        new_row[month + '_Details'] = ''
                        new_row[month + '_Date'] = ''
                    df_excel = pd.concat([df_excel, pd.DataFrame([new_row])], ignore_index=True)

            df_excel.to_excel(writer, sheet_name=year, index=False, header=True)

    save_excel_to_github(excel_buffer)

temp_storage = []

def model_inference(transaction_details, formatted_date):
    inputs = tokenizer(transaction_details, return_tensors="pt", truncation=True, padding=True)
    with torch.no_grad():
        outputs = model(**inputs)
       
    logits = outputs.logits
    predicted_label_ids = torch.argmax(logits, dim=2)
    label_list = ["O", "B-NAME", "I-NAME", "B-SENT_TO", "I-SENT_TO"]

    predictions = [label_list[label_id] for label_id in predicted_label_ids[0].cpu().numpy()]
    
    tokens = tokenizer.convert_ids_to_tokens(inputs["input_ids"][0])
    word_pieces = []
    current_label = "O"
    collected_sent_to = []

    for token, prediction in zip(tokens, predictions):
        if token.startswith("##"):
            word_pieces.append(token[2:])
        else:
            if word_pieces and current_label == "B-SENT_TO":
                collected_sent_to.append(''.join(word_pieces))
            
            if prediction == "B-SENT_TO":
                word_pieces = [token]
            elif prediction == "I-SENT_TO" and word_pieces:
                word_pieces.append(token)
            else:
                word_pieces = []

            current_label = prediction

    if word_pieces and current_label == "B-SENT_TO":
        collected_sent_to.append(''.join(word_pieces))
    
    group_category = "Others"
    iteration = collected_sent_to
   
    mobile = ["AirtelPostpaid", "jiomobile", "airtel"]
    mobile_lower = [m.lower() for m in mobile]

    for collected_sent in iteration:
        collected_sent_to_lower = collected_sent.lower()
        if any(m in collected_sent_to_lower for m in mobile_lower):
            group_category = "Mobile"
            break

    result = {
        "Group": group_category,
        "details": collected_sent_to,
        "date": formatted_date,
    }
    return result



@app.route('/endpoint', methods=['POST'])
def endpoint():
    if request.method == 'POST':
        data = request.get_json()
        transaction_details = data.get('transactionDetails')
        formatted_date = data.get('formattedDate')

        # Perform inference using the model
        result = model_inference(transaction_details, formatted_date)
        
        try:
            save_to_excel([result])  # Wrap the single result in a list
        except Exception as e:
            print(f"Failed to update Excel: {e}")
        
        return jsonify(result)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
